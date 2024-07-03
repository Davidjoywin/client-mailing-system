import openpyxl
import openpyxl.workbook

from google_sheet import googleSheet

wb = openpyxl.load_workbook("sample.xlsx")

ws = wb.active

def get_content():
    outer = []
    for row in range(1, ws.max_row+1):
        inner = []
        for column in range(1, ws.max_column+1):
            item = ws.cell(row=row, column=column).value
            inner.append(item)
        outer.append(inner)
    return outer

def get_key(key):
    for i in range(1, ws.max_column+1):
        cell = ws.cell(row=1, column=i).value
        if cell.lower() == key:
            return i
    return 0

def list_cell_title():
    for i in range(1, ws.max_column+1):
        cell = ws.cell(row=1, column=i)
        yield cell.value

def get_column_values(title):
    key = get_key(title)
    for row in range(1, ws.max_column+1):
        cell = ws.cell(row=row, column=key)
        yield cell.value

def get_row_values(value):
    for column in get_content():
        for row in column:
            if row.lower() == value:
                return column
    return None

# for value in get_column_values("email address"):
#     print(value)

# for item in get_content():
#     for it in item:
#         print(it, end=' ')
#     print()

# key = get_key("email address")
# print(key)

# for title in list_cell_title():
#     print(title)

# value = get_row_values("email")
# print(value)